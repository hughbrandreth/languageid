import csv
import numpy
from openpyxl import Workbook
from tkinter import *

from aug22.source.Timelogger import TimeLogsDict
from aug22.source.framework.Comparison import Comparison
from aug22.source.framework.Language import Language
from aug22.source.framework.Text import Text
from aug22.source.framework.utilities import FullLangList
from aug22.source.resultsprocessing.Table import Table


class GrandAnalysis:

    def __init__(self, TextFileNameList, LangNameList, AlgorithmNameList):

        self.SuccessfullyInitialised = True
        for InputList in [TextFileNameList, LangNameList, AlgorithmNameList]:
            if type(InputList) != list:
                self.SuccessfullyInitialised = False
            elif len(InputList) == 0:
                self.SuccessfullyInitialised = False

        if self.SuccessfullyInitialised:

            self.AlgorithmNameList = AlgorithmNameList
            self.TextFileNameList = TextFileNameList
            self.TextIndexDict = dict()
            self.TextCommentList = []
            self.TextList = []
            self.LangNameList = LangNameList

            for TextFileName in self.TextFileNameList:
                File = open(r"E:/Hugh/HDD Documents/Python Programs/Decipheronator/source/sampletexts/" + TextFileName, "r", encoding="utf8")
                for Line in File.readlines():
                    if Line[0] == "#":
                        self.TextCommentList.append(Line.strip()) ####### not needed -- maybe suppress #####
                    else:
                        ThisText = Text(Line.strip(), TextFileName)
                        if ThisText.Lang not in self.LangNameList and ThisText.Lang in FullLangList:
                            LangNameList.append(ThisText.Lang)
                        self.TextList.append(ThisText)
                File.close()




            self.LangCount = len(self.LangNameList)
            self.TextCount = len(self.TextList)
            self.AlgorithmCount = len(self.AlgorithmNameList)

            self.BestScoreListDict = dict()
            self.BestScoringLangListDict = dict()
            self.AlgorithmTimeDict = dict()

            self.GenerateComparisons()
            self.GenerateLangConfusionMatricesAndAlgorithmAccuracyScores()
            self.TextContentList = list(map(lambda t: t.Content, self.TextList))

        return

    ##### INTERNAL FUNCTIONS #####

    def CheckInitiation(self):
        return self.SuccessfullyInitialised

    def GenerateComparisons(self):
        '''
        Constructs a lang object for every requested language, and constructs a comparison object for each language and text.
        Performs every requested algorithm on each comparison.
        Results of each algorithm are stored in a numpy array.
        The highest value and language which scored the highest value on every algorithm with every text is computed and stored in two more numpy arrays.

        :return: None
        '''
        if not self.SuccessfullyInitialised:
            return

        self.ScoreArray = numpy.empty((self.LangCount, self.AlgorithmCount, self.TextCount))
        for LangIndex, LangName in enumerate(self.LangNameList):
            Lang = Language("wordfreq", LangName)
            Lang.FindLetterFreqs()          #### Initialisation seems to take a lot of time and makes the algorithm seem very slow, so we are trying to separate initialisation from the running.
            for AlgorithmIndex, AlgorithmName in enumerate(self.AlgorithmNameList):

                for TextIndex, Text in enumerate(self.TextList):
                    ThisComparison = Comparison(Text, Lang)
                    Algorithm = getattr(ThisComparison, AlgorithmName)
                    self.ScoreArray[LangIndex, AlgorithmIndex, TextIndex] = Algorithm()
                self.AlgorithmTimeDict[AlgorithmName] = TimeLogsDict[AlgorithmName]
        self.BestScoreArray = numpy.amax(self.ScoreArray, axis=0)
        self.BestScoreIndexArray = numpy.argmax(self.ScoreArray, axis=0)
        self.BestScoringLangArray = numpy.vectorize(lambda i: self.LangNameList[i])(self.BestScoreIndexArray)

        self.MeanScoreArray = numpy.mean(self.ScoreArray, axis=1)
        self.BestMeanScoreArray = numpy.amax(self.MeanScoreArray, axis=0)
        self.BestMeanScoreIndexArray = numpy.argmax(self.MeanScoreArray, axis=0)
        self.BestMeanScoringLangArray = numpy.vectorize(lambda i: self.LangNameList[i])(self.BestMeanScoreIndexArray)
        return

    def GenerateLangConfusionMatricesAndAlgorithmAccuracyScores(self):
        '''
        Takes no parameters, but compares the language a text was *actually* in (which is read from its source file) from the language that each algorithm estimated that it was in.
        An estimation means the language which the algorithm gave the highest score to.
        This is used to calculate a success rate for each algorithm on each language.
        The success rates are stored in another numpy array in the object.

        :return:
        '''

        if not self.SuccessfullyInitialised:
            return

        self.AlgorithmSuccessCountList = []
        self.LangConfusionMatrixArray = numpy.zeros((len(self.AlgorithmNameList), len(self.LangNameList), len(self.LangNameList)))
        for AlgorithmIndex, AlgorithmName in enumerate(self.AlgorithmNameList):
            SuccessCount = 0
            for TextIndex, Text in enumerate(self.TextList):
                if Text.Lang in self.LangNameList:
                    EstimatedLangIndex = self.BestScoreIndexArray[AlgorithmIndex, TextIndex]
                    ActualLangIndex = self.LangNameList.index(Text.Lang)
                    self.LangConfusionMatrixArray[AlgorithmIndex, EstimatedLangIndex, ActualLangIndex] += 1
                    if EstimatedLangIndex == ActualLangIndex:
                        SuccessCount += 1
            self.AlgorithmSuccessCountList.append(SuccessCount)
        return


    ########   OUTPUT Functions   ########

    def GetBestScore(self, Algorithm, TextListIndex):
        '''
        Retrieves the best score an algorithm gave to any language on a given text.

        :param str Algorithm:
        :param int TextListIndex:
        :return: float between 0 and 1
        '''
        if not self.SuccessfullyInitialised:
            return None
        else:
            return self.BestScoreArray[self.AlgorithmNameList.index(Algorithm), TextListIndex]

    def GetBestScoringLang(self, Algorithm, TextListIndex):
        '''
        Retrieves the name (code) of the language which received the highest score from the given algorithm on the given text.

        :param str Algorithm:
        :param int TextListIndex:
        :return: language code as string
        '''

        if not self.SuccessfullyInitialised:
            return "??"
        else:
            return self.BestScoringLangArray[self.AlgorithmNameList.index(Algorithm),TextListIndex]

    ###### CSV RELATED OPERATIONS ######

    def SaveAsCsv(self, FileName):
        '''
        Saves the results of the analysis in a csv in the file location provided.
        The csv is opened and a chart of the results for every text recorded.
        Only records the results of the first algorithm.

        :param str FileName:
        :return: None
        '''

        if not self.SuccessfullyInitialised:
            return

        File = open(FileName, "w", newline = "", encoding="utf8")
        Writer = csv.writer(File)

        Writer.writerow(["The most likely language is:", self.BestScoringLangArray[0, 0]])
        BestScoreRow = [str("Score of " + str(self.BestScoringLangArray[0, 0]) + " was:"), self.BestScoreArray[0, 0]]
        Writer.writerow(BestScoreRow)
        Writer.writerow("")
        Writer.writerow(["Full list of Scores:"])
        Header = self.LangNameList.copy()
        Header.insert(0, "Text:")
        Header.insert(1, "Source:")
        Writer.writerow(Header)
        for TextIndex, Text in enumerate(self.TextList):
            Row = [Text.Content, Text.Source]
            for LangIndex, Lang in enumerate(self.LangNameList):
                Row.append(self.ScoreArray[LangIndex, 0, TextIndex])
            Writer.writerow(Row)

        return

    ###### EXCEL RELATED OPERATIONS ######

    def SaveAsExcel(self, FileName):
        '''
        This will create an excel Spreadsheet in the location specified and populate it with all of the results.
        A sheet will be created summarising the performance of all of the algorithms.
        A sheet will be created for each algorithm describing how it scored every language on every text, and a breakdown of its success rate and time taken.

        :param str FileName:
        :return: None
        '''

        if not self.SuccessfullyInitialised:
            return

        WorkBook = Workbook()

        SummaryWorkSheet = WorkBook.create_sheet("Summary")
        SummaryWorkSheet.cell(1, 1).value = "Algorithm"
        SummaryWorkSheet.cell(1, 2).value = "Success Rate"
        SummaryWorkSheet.cell(1, 3).value = "Average Time Taken Per Text"

        for AlgorithmIndex, (AlgorithmName, AlgorithmSuccessCount) in enumerate(zip(self.AlgorithmNameList, self.AlgorithmSuccessCountList)):

            WorkSheet = WorkBook.create_sheet(AlgorithmName)
            NextRow = 1
            WorkSheet.cell(NextRow, 1).value = "Text:"
            WorkSheet.cell(NextRow, 2).value = "Source:"


            ## SCORES Array HEADER
            for LangNameIndex, LangName in enumerate(self.LangNameList):
                WorkSheet.cell(NextRow, 3 + LangNameIndex).value = LangName + " Score"
            HighScoreColumn = len(self.LangNameList) + 3
            WorkSheet.cell(NextRow, HighScoreColumn).value = "Highest Score:"
            BestScoringLangColumn = HighScoreColumn + 1
            WorkSheet.cell(NextRow, BestScoringLangColumn).value = "Most Likely Language:"
            NextRow += 1

            ##  SCORES Array
            for TextIndex, Text in enumerate(self.TextList):
                WorkSheet.cell(NextRow + TextIndex, 1).value = Text.Content
                WorkSheet.cell(NextRow + TextIndex, 2).value = Text.Source
                for LangIndex, Lang in enumerate(self.LangNameList):
                    WorkSheet.cell(NextRow + TextIndex, 3 + LangIndex).value = self.ScoreArray[LangIndex, AlgorithmIndex, TextIndex]
                WorkSheet.cell(NextRow + TextIndex, HighScoreColumn).value = self.BestScoreArray[AlgorithmIndex, TextIndex]
                WorkSheet.cell(NextRow + TextIndex, BestScoringLangColumn).value = self.BestScoringLangArray[AlgorithmIndex, TextIndex]
            NextRow += self.TextCount + 1

            ## CONFUSION MATRIX
            # This is a martix comparing the languages that texts were *actually* in to the languages that the algorithm predicted that they were in #
            Langtable = Table(self.LangNameList, self.LangNameList, numpy.transpose(self.LangConfusionMatrixArray[AlgorithmIndex, :, :]), "Language Confusion Matrix")
            Langtable.outputtoxlsx(WorkSheet, NextRow, 1)
            NextRow += len(self.LangNameList) + 3

            WorkSheet.cell(NextRow, 1).value = "Total Number of Texts Analysed: "
            WorkSheet.cell(NextRow, 2).value = len(self.TextList)
            NextRow += 1
            WorkSheet.cell(NextRow, 1).value = "Total Number Texts Correctly Identified: "
            WorkSheet.cell(NextRow, 2).value = AlgorithmSuccessCount
            NextRow += 1
            WorkSheet.cell(NextRow, 1).value = "Proportion of Texts Correctly Identified: "
            WorkSheet.cell(NextRow, 2).value = AlgorithmSuccessCount / len(self.TextList)

            NextRow += 2
            WorkSheet.cell(NextRow, 1).value = "Data Sourced from the Following Files:"
            NextRow += 1
            for Row, TextSource in enumerate(self.TextFileNameList):
                WorkSheet.cell(NextRow + Row, 1).value = TextSource
            NextRow += len(self.TextFileNameList) + 2

            WorkSheet.cell(NextRow, 1).value = "Total Runtime of " + AlgorithmName
            NextRow += 1
            WorkSheet.cell(NextRow, 1).value = sum(self.AlgorithmTimeDict[AlgorithmName])
            NextRow += 1
            WorkSheet.cell(NextRow, 1).value = "Average Runtime of " + AlgorithmName + ":"
            NextRow += 1
            WorkSheet.cell(NextRow, 1).value = sum(self.AlgorithmTimeDict[AlgorithmName]) / len(self.AlgorithmTimeDict[AlgorithmName])

            SummaryWorkSheet.cell(AlgorithmIndex + 3, 1).value = AlgorithmName
            SummaryWorkSheet.cell(AlgorithmIndex + 3, 2).value = AlgorithmSuccessCount / len(self.TextList)
            SummaryWorkSheet.cell(AlgorithmIndex + 3, 3).value = sum(self.AlgorithmTimeDict[AlgorithmName]) / len(self.AlgorithmTimeDict[AlgorithmName])


#### For the Mean Algorithm

        WorkSheet = WorkBook.create_sheet("Mean of All Algorithms")
        NextRow = 1
        WorkSheet.cell(NextRow, 1).value = "Text:"
        WorkSheet.cell(NextRow, 2).value = "Source:"

        ## SCORES Array HEADER
        for LangNameIndex, LangName in enumerate(self.LangNameList):
            WorkSheet.cell(NextRow, 3 + LangNameIndex).value = LangName + " Score"
        HighScoreColumn = len(self.LangNameList) + 3
        WorkSheet.cell(NextRow, HighScoreColumn).value = "Highest Score:"
        BestScoringLangColumn = HighScoreColumn + 1
        WorkSheet.cell(NextRow, BestScoringLangColumn).value = "Most Likely Language:"
        NextRow += 1

        ##  SCORES Array
        for TextIndex, Text in enumerate(self.TextList):
            WorkSheet.cell(NextRow + TextIndex, 1).value = Text.Content
            WorkSheet.cell(NextRow + TextIndex, 2).value = Text.Source
            for LangIndex, Lang in enumerate(self.LangNameList):
                WorkSheet.cell(NextRow + TextIndex, 3 + LangIndex).value = self.MeanScoreArray[
                    LangIndex, TextIndex]
            WorkSheet.cell(NextRow + TextIndex, HighScoreColumn).value = self.BestMeanScoreArray[TextIndex]
            WorkSheet.cell(NextRow + TextIndex, BestScoringLangColumn).value = self.BestMeanScoringLangArray[TextIndex]

        NextRow += self.TextCount + 2
        WorkSheet.cell(NextRow, 1).value = "Data Sourced from the Following Files:"
        NextRow += 1
        for Row, TextSource in enumerate(self.TextFileNameList):
            WorkSheet.cell(NextRow + Row, 1).value = TextSource
        NextRow += len(self.TextFileNameList) + 2


        WorkBook.remove_sheet(WorkBook.get_sheet_by_name("Sheet"))
        WorkBook.save(filename=FileName)

        return

    ###### TEXT  RELATED OPERATIONS ######

    def SaveAsText(self, FileName):
        '''
        Creates a text file in the specified location and writes the score of every language on every text analysed.
        Only does so for the first algorithm provided.

        :param str FileName:
        :return: None
        '''

        if not self.SuccessfullyInitialised:
            return

        File = open(FileName, "w", encoding="utf8")

        File.write("The most likely language is:" + str(self.BestScoringLangArray[0, 0]) + "\n")
        File.write("Score of " + str(self.BestScoringLangArray[0, 0]) + " was:" + str(self.BestScoreArray[0, 0]) + "\n")
        File.write("Full list of Scores:" + "\n")


        File.write("Text:                | Source:              | " + '               | '.join(self.LangNameList) + "\n")
        for TextIndex, Text in enumerate(self.TextList):
            Row = []
            for LangIndex, Lang in enumerate(self.LangNameList):
                Row.append(str(self.ScoreArray[LangIndex, 0, TextIndex]))
            File.write(Text.Content[:10] + "... | " + Text.Source + " | " + ' |'.join(Row) + "\n")



