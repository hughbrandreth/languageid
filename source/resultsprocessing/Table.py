import csv
from openpyxl import Workbook

from openpyxl.styles import Font, Color
import numpy


class Table:
    
    def __init__(self, RowNames, ColumnNames, Content, title=None):
        self.RowNames = RowNames
        self.ColumnNames = ColumnNames
        self.Content = Content.tolist()
        self.title = title
        self.Height = len(self.RowNames)
        self.Width = len(self.ColumnNames)
        self.Size = (self.Height, self.Width)

        
    def asList(self):

        tableasList = [[self.title], self.ColumnNames.copy()]
        tableasList[1].insert(0, "")
        for RowName, ContentRow in zip(self.RowNames, self.Content):
            Row = ContentRow.copy()
            Row.insert(0, RowName)
            tableasList.append(Row)

        return tableasList


    def outputtocsv(self, fileName):
        file = open(fileName, "w", newline="")
        writer = csv.writer(file)
        for Row in self.asList():
            writer.writerow(Row)
        file.close()
        return


    def outputtoxlsx(self, WorkSheet, Row, Column):

        WorkSheet.cell(Row, Column).value = self.title
        for c, ColumnName in enumerate(self.ColumnNames):
            WorkSheet.cell(Row + 1, Column + c + 1).value = ColumnName

        for r, (RowName, ContentRow) in enumerate(zip(self.RowNames, self.Content)):
            WorkSheet.cell(Row + r + 2, Column).value = RowName
            for c, (cellvalue) in enumerate(ContentRow):
                WorkSheet.cell(Row + r + 2, Column + c + 1).value = cellvalueS


        return